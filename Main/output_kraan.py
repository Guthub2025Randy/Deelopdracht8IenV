# -*- coding: utf-8 -*-
"""
Created on Fri Mar  7 09:31:01 2025

@author: CWMaz
"""
from openpyxl import load_workbook

def output_kraan(SWLmax,gewichttransitionpiece,lengte_kraan_fundatie,Draaihoogte_kraan,jib_length,Zwenkhoek,Giekhoek,LCG_TP,TCG_TP,VCG_TP,LCG_kraanhuis,TCG_kraanhuis,VCG_kraanhuis,LCG_kraanboom,TCG_kraanboom,VCG_kraanboom,LCG_heisgerei,TCG_heisgerei,VCG_heisgerei):
   
    bestandspad= "Voorbeeldschip_Deelopdracht_8_V3.xlsx"

    wb= load_workbook(bestandspad)

    ws=wb.active 
    
    ws["D65"] = SWLmax # SWLmax van de kraan
    ws["D67"] = gewichttransitionpiece #gewicht per transition piece
    ws["D69"] = lengte_kraan_fundatie #lengte_kraanfundatie 
    ws["D70"] = Draaihoogte_kraan #Draaihoogte_kraan
    ws["D71"] = jib_length# kraanboom lengte
    ws["D72"] = Zwenkhoek #zwenkhoek
    ws["D73"] = Giekhoek #giekhoek
    ws["D75"] = LCG_TP #lcg van de kraanlast
    ws["D76"] = TCG_TP #tcg van de kraanlast
    ws["D77"] = VCG_TP #vcg van de kraanlast
    ws["D79"] = LCG_kraanhuis #LCG_kraanhuis
    ws["D80"] = TCG_kraanhuis #TCG_kraanhuis
    ws["D81"] = VCG_kraanhuis #VCG_kraanhuis
    ws["D83"] = LCG_kraanboom #LCG_kraanboom
    ws["D84"] = TCG_kraanboom #TCG_kraanboom
    ws["D85"] = VCG_kraanboom #VCG_kraanboom
    ws["D87"] = LCG_heisgerei #LCG_heisgerei
    ws["D88"] = TCG_heisgerei #VCG_heisgerei
    ws["D89"] = VCG_heisgerei #TCG_heisgerei
    
    wb.save(bestandspad)
    return None

output_kraan(1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19)