# -*- coding: utf-8 -*-
"""
Created on Wed May  7 11:08:35 2025

@author: CWMaz
"""

from openpyxl import load_workbook
from bibliotheek import *

def output_kraan(SWLmax,gewichttransitionpiece,lengte_kraan_fundatie,Draaihoogte_kraan,jib_length,Zwenkhoek,Giekhoek,LCG_TP,TCG_TP,VCG_TP,LCG_kraanhuis,TCG_kraanhuis,VCG_kraanhuis,LCG_kraanboom,TCG_kraanboom,VCG_kraanboom,LCG_heisgerei,TCG_heisgerei,VCG_heisgerei, versienummer):

    """Deze functie regelt de output voor alles wat met de kraan te maken heeft en zet het op de juiste plek in het antwoordenblad"""
   
    bestandspad= "Antwoordenblad_MT1463_1466_V0{0}.xlsx".format(versienummer)

    wb= load_workbook(bestandspad)

    ws=wb.active 
    
    ws["D65"] = SWLmax # SWLmax van de kraan
    ws["D67"] = gewichttransitionpiece #gewicht per transition piece
    ws["D69"] = lengte_kraan_fundatie #lengte_kraanfundatie 
    ws["D70"] = Draaihoogte_kraan #Draaihoogte_kraan
    ws["D71"] = jib_length# kraanboom lengte
    ws["D72"] = Zwenkhoek #zwenkhoek
    ws["D73"] = Giekhoek #giekhoek
    ws["D75"] = LCG_heisgerei #lcg van de kraanlast
    ws["D76"] = TCG_heisgerei #tcg van de kraanlast
    ws["D77"] = VCG_heisgerei #vcg van de kraanlast
    ws["D79"] = LCG_kraanhuis #LCG_kraanhuis
    ws["D80"] = TCG_kraanhuis #TCG_kraanhuis
    ws["D81"] = VCG_kraanhuis #VCG_kraanhuis
    ws["D83"] = LCG_kraanboom #LCG_kraanboom
    ws["D84"] = TCG_kraanboom #TCG_kraanboom
    ws["D85"] = VCG_kraanboom #VCG_kraanboom
    ws["D87"] = LCG_heisgerei #LCG_heisgerei
    ws["D88"] = TCG_heisgerei #VCG_heisgerei
    ws["D89"] = VCG_heisgerei #TCG_heisgerei
    
    wb.save(bestandspad)
    return None

def output_1(version,entrance_angle_location,R_14knp,gm,dikte_huid_en_dek,loa,B,H,T,trim,heel,dichtheid_staal,dichtheid_water,Deplacement,L_C_G,T_C_G,V_C_G,opdrijvendekracht,L_C_B,T_C_B,V_C_B,DVF,DLM,DTM,aantal_transition_pieces,massa_transition_pieces,lcg_tp,tcg_tp,vcg_tp, versienummer):
    
    """deze functie regelt een deel van de output en zorgt er voor dat het op de juiste plek in het excel document komt te staan."""
    
    bestandspad= "Antwoordenblad_MT1463_1466_V0{0}.xlsx".format(versienummer)
    wb= load_workbook(bestandspad)
    ws=wb.active 
    ws["D8"]  = version  #versie 
    ws["D11"] = entrance_angle_location#entrance angle location 
    ws["D12"] = R_14knp #totale weerstand bij 14 knoop
    ws["D16"] = gm #GM
    ws["D19"] = dikte_huid_en_dek #dikte huid en dek
    ws["D22"] = loa #lengte
    ws["D23"] = B #breedte
    ws["D24"] = H #holte
    ws["D25"] = T #diepgang
    ws["D26"] = trim #trim
    ws["D27"] = heel #helling
    ws["D31"] = dichtheid_staal # dichtheid staal
    ws["D32"] = dichtheid_water #dichtheid water
    ws["D36"] = Deplacement # deplacement
    ws["D37"] = L_C_G #LCG(x)
    ws["D38"] = T_C_G #TCG(y)
    ws["D39"] = V_C_G #VCG (z)
    ws["D43"] = opdrijvendekracht # opdrijvende kracht
    ws["D44"] = L_C_B #LCB
    ws["D45"] = T_C_B #TCB
    ws["D46"] = V_C_B #VCB
    ws["D49"] = DVF#Afwijking van verticaal krachtevenwicht
    ws["D50"] = DLM#Afwijking van longitudinaal momentevenwicht
    ws["D51"] = DTM#Afwijking van transversaal momentevenwicht
    ws["D55"] = aantal_transition_pieces
    ws["D56"] = massa_transition_pieces
    ws["D57"] = lcg_tp
    ws["D58"] = tcg_tp
    ws["D59"] = vcg_tp
    wb.save(bestandspad)
    return None

def output_globale_sterkte(Maximaal_moment, LMmilvda, Maximale_afschuiving, LMailvda, Maximale_doorbuiging, LMdilvda):
    
    """deze functie regelt de output van de globale sterkte waarden en zet ze op de juiste plek in de excel sheet"""
    
    bestandspad="Antwoordenblad_MT1463_1466_V{0}.xlsx".format(versienummers)
    wb= load_workbook(bestandspad)
    ws=wb.active
    ws["D92"]=Maximaal_moment 
    ws["D93"]=LMmilvda #Locatie Maximaal moment in lengte vanaf de achterloodlijn
    ws["D94"]=Maximale_afschuiving
    ws["D95"]=LMailvda #Locatie Maximale afschuiving in lengte vanaf de achterloodlijn
    ws["D96"]=Maximale_doorbuiging
    ws["D97"]=LMdilvda #Locatie Maximale doorbuiging in lengte vanaf de achterloodlijn
    return None

